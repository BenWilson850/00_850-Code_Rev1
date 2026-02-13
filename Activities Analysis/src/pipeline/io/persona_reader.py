"""Readers for the persona (client workbook) Excel format."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

import openpyxl

from ..core.persona_logic import clean_number
from ..core.persona_match import infer_test_id
from ..core.persona_models import ClientInfo
from ..utils import get_logger


_log = get_logger(__name__)


@dataclass(frozen=True)
class ClientSheetData:
    client: ClientInfo
    # test_id -> {"label": str, "5": float|None, "10": float|None}
    tests: dict[str, dict[str, Any]]


def read_persona_workbook(
    path: str | Path,
    client_config: dict[str, Any] | None = None,
) -> list[ClientSheetData]:
    cfg = client_config or {}
    metadata_rows = int(cfg.get("metadata_rows", 4))
    table_start_row = int(cfg.get("table_start_row", 5))
    test_col = str(cfg.get("test_col", "A")).upper()
    yr5_col = str(cfg.get("yr5_col", "C")).upper()
    yr10_col = str(cfg.get("yr10_col", "D")).upper()

    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Client workbook not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out: list[ClientSheetData] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        meta: dict[str, Any] = {}
        for r in range(1, metadata_rows + 1):
            k = ws[f"A{r}"].value
            v = ws[f"B{r}"].value
            if k is None and v is None:
                continue
            key = str(k).strip().lower() if k is not None else ""
            if key:
                meta[key] = v

        name = str(meta.get("name") or "").strip() or str(sheet_name).strip()
        age_raw = meta.get("age")
        age_cleaned = clean_number(age_raw)
        age = int(age_cleaned) if age_cleaned is not None else None

        gender_raw = meta.get("gender")
        gender: Literal["M", "F"] | None = None
        if gender_raw is not None:
            gs = str(gender_raw).strip().lower()
            if gs in {"m", "male"} or "male" in gs:
                gender = "M"
            elif gs in {"f", "female"} or "female" in gs:
                gender = "F"

        client = ClientInfo(name=name, age=age, gender=gender, sheet_name=sheet_name)

        tests: dict[str, dict[str, Any]] = {}
        row = table_start_row
        # Read until we hit a run of blank test names.
        blank_seen = 0
        while row <= ws.max_row and blank_seen < 10:
            test_label = ws[f"{test_col}{row}"].value
            if test_label is None or str(test_label).strip() == "":
                blank_seen += 1
                row += 1
                continue
            blank_seen = 0

            test_id = infer_test_id(test_label)
            v5 = clean_number(ws[f"{yr5_col}{row}"].value)
            v10 = clean_number(ws[f"{yr10_col}{row}"].value)
            tests[test_id] = {"label": str(test_label).strip(), "5": v5, "10": v10}
            row += 1

        if not tests:
            _log.warning("No tests parsed for sheet '%s' in %s", sheet_name, path)

        out.append(ClientSheetData(client=client, tests=tests))

    _log.info("Read %s client sheets from %s", len(out), path)
    return out

