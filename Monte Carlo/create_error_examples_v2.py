"""Create error examples for demonstration (v2 - corrected structure)."""

import openpyxl
from pathlib import Path


def introduce_errors(file_path: Path):
    """Introduce various types of errors."""
    wb = openpyxl.load_workbook(file_path)

    print("\n" + "="*70)
    print("INTRODUCING TEST ERRORS")
    print("="*70)

    sheet_names = [s for s in wb.sheetnames][:4]  # First 4 clients

    # ERROR 1: Wrong direction - Improvement on decline
    sheet1 = wb[sheet_names[0]]
    for row_idx in range(2, sheet1.max_row + 1):
        test_name = sheet1.cell(row_idx, 1).value
        if test_name == 'VO2 max':
            baseline = sheet1.cell(row_idx, 2).value
            if baseline:
                # Make Year 5 HIGHER (improvement on decline - WRONG!)
                improved_value = baseline * 1.12
                sheet1.cell(row_idx, 3).value = improved_value
                print(f"\n[X] ERROR 1 - {sheet_names[0]}: WRONG DIRECTION")
                print(f"  Test: VO2 max (higher is better)")
                print(f"  Scenario: Decline (should worsen)")
                print(f"  Baseline: {baseline:.1f}")
                print(f"  Year 5: {improved_value:.1f}  IMPROVED instead of declined!")
            break

    # ERROR 2: Non-monotonic progression
    sheet2 = wb[sheet_names[1]]
    for row_idx in range(2, sheet2.max_row + 1):
        test_name = sheet2.cell(row_idx, 1).value
        if test_name == 'Grip Strength':
            baseline = sheet2.cell(row_idx, 2).value
            if baseline:
                # Year 5 declines, then Year 10 REVERSES (goes back up - WRONG!)
                year5 = baseline * 0.88
                year10 = baseline * 0.91  # Goes back UP
                sheet2.cell(row_idx, 3).value = year5
                sheet2.cell(row_idx, 4).value = year10
                print(f"\n[X] ERROR 2 - {sheet_names[1]}: NON-MONOTONIC PROGRESSION")
                print(f"  Test: Grip Strength (higher is better)")
                print(f"  Scenario: Decline (should decrease monotonically)")
                print(f"  Baseline: {baseline:.1f}")
                print(f"  Year 5: {year5:.1f}  Declined")
                print(f"  Year 10: {year10:.1f}  REVERSED (went back up!)")
            break

    # ERROR 3: Physiological implausibility - Too high
    sheet3 = wb[sheet_names[2]]
    for row_idx in range(2, sheet3.max_row + 1):
        test_name = sheet3.cell(row_idx, 1).value
        if test_name == 'Body Fat %':
            # Set to impossible value
            impossible_value = 72.0
            sheet3.cell(row_idx, 4).value = impossible_value
            print(f"\n[X] ERROR 3 - {sheet_names[2]}: PHYSIOLOGICAL IMPLAUSIBILITY")
            print(f"  Test: Body Fat %")
            print(f"  Year 10: {impossible_value:.1f}%")
            print(f"  Problem: Humans cannot have >60% body fat (survival limit)")
            break

    # ERROR 4: Physiological implausibility - Negative value
    sheet4 = wb[sheet_names[3]]
    for row_idx in range(2, sheet4.max_row + 1):
        test_name = sheet4.cell(row_idx, 1).value
        if test_name == 'Gait Speed':
            impossible_value = -0.3
            sheet4.cell(row_idx, 4).value = impossible_value
            print(f"\n[X] ERROR 4 - {sheet_names[3]}: PHYSIOLOGICAL IMPLAUSIBILITY")
            print(f"  Test: Gait Speed")
            print(f"  Year 10: {impossible_value:.1f} m/s")
            print(f"  Problem: Negative gait speed is physically impossible")
            break

    print("\n" + "="*70)
    print("Saving corrupted test file...")
    wb.save(file_path)
    print(f"[OK] Saved: {file_path.name}")
    print("="*70)


def main():
    base_path = Path(__file__).parent
    test_file = base_path / "outputs" / "MonteCarlo_Decline_TEST.xlsx"

    if not test_file.exists():
        print(f"Error: {test_file} not found")
        return

    introduce_errors(test_file)

    print("\n" + "="*70)
    print("NEXT STEP: Run validation to see how these errors are detected")
    print("="*70)


if __name__ == "__main__":
    main()
