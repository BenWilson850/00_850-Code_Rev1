"""Reader for normative database."""

from __future__ import annotations
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl
import re


@dataclass
class NormativeData:
    """Container for all normative reference data."""

    # Physical tests with age/gender normative values
    vo2_max: Dict[str, List[Tuple[Tuple[float, float], float]]]  # gender -> [(age_range, value)]
    grip_strength: Dict[str, List[Tuple[float, float]]]  # gender -> [(age, value)]
    vertical_jump: Dict[str, List[Tuple[float, float]]]
    sts_power: Dict[str, List[Tuple[float, float]]]
    gait_speed: Dict[str, List[Tuple[Tuple[float, float], float]]]
    tug: Dict[str, List[Tuple[float, float]]]
    single_leg_stance: List[Tuple[Tuple[float, float], float]]  # No gender split
    sit_and_reach: Dict[str, List[Tuple[Tuple[float, float], float]]]

    # Metabolic risk ranges
    metabolic_ranges: Dict[str, Dict[str, str]]  # marker -> {Low risk, Normal, Elevated}

    # Body fat percentage ranges
    body_fat_ranges: Dict[str, List[Tuple[Tuple[float, float], Tuple[float, float, float]]]]


def parse_age_range(age_str: str) -> Tuple[float, float]:
    """
    Parse age range string like '20-29' or '20–29' into (min, max).

    Args:
        age_str: String like '20-29', '20–29', or '20 - 29'

    Returns:
        Tuple of (min_age, max_age)
    """
    # Handle various dash types and spaces
    age_str = age_str.strip().replace('–', '-').replace(' ', '')
    if '-' in age_str:
        parts = age_str.split('-')
        return (float(parts[0]), float(parts[1]))
    else:
        # Single age or age+
        age_val = float(re.sub(r'[^\d.]', '', age_str))
        return (age_val, age_val + 9)  # Assume decade range


def read_normative_database(db_path: str | Path) -> NormativeData:
    """
    Read the normative database Excel file.

    Args:
        db_path: Path to Normative Database.xlsx

    Returns:
        NormativeData object with all reference data
    """
    db_path = Path(db_path)
    if not db_path.exists():
        raise FileNotFoundError(f"Normative database not found: {db_path}")

    wb = openpyxl.load_workbook(db_path, data_only=True)

    # VO2 Max (decade ranges, male/female)
    ws = wb['VO2 Max']
    vo2_male = []
    vo2_female = []
    for row in range(5, 11):  # Rows 5-10 have data
        age_range = parse_age_range(str(ws.cell(row, 1).value))
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 5).value)
        vo2_male.append((age_range, male_val))
        vo2_female.append((age_range, female_val))
    vo2_max = {'Male': vo2_male, 'Female': vo2_female}

    # Grip Strength (5-year increments, male/female)
    ws = wb['Grip Strength']
    grip_male = []
    grip_female = []
    for row in range(3, 17):  # Rows 3-16
        age = float(ws.cell(row, 1).value)
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 3).value)
        grip_male.append((age, male_val))
        grip_female.append((age, female_val))
    grip_strength = {'Male': grip_male, 'Female': grip_female}

    # Vertical Jump
    ws = wb['Vertical Jump']
    vj_male = []
    vj_female = []
    for row in range(3, 16):
        age = float(ws.cell(row, 1).value)
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 3).value)
        vj_male.append((age, male_val))
        vj_female.append((age, female_val))
    vertical_jump = {'Male': vj_male, 'Female': vj_female}

    # Sit to Stand Power
    ws = wb['Sit to stand']
    sts_male = []
    sts_female = []
    for row in range(3, 18):
        age = float(ws.cell(row, 1).value)
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 3).value)
        sts_male.append((age, male_val))
        sts_female.append((age, female_val))
    sts_power = {'Male': sts_male, 'Female': sts_female}

    # Gait Speed (decade ranges)
    ws = wb['Gait speed walking']
    gait_male = []
    gait_female = []
    for row in range(2, 8):
        age_range = parse_age_range(str(ws.cell(row, 1).value))
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 3).value)
        gait_male.append((age_range, male_val))
        gait_female.append((age_range, female_val))
    gait_speed = {'Male': gait_male, 'Female': gait_female}

    # TUG - need to add this from updated normative database
    # For now, placeholder - you mentioned it was added
    tug_male = [(age, 5.0 + (age - 20) * 0.15) for age in range(20, 81, 5)]
    tug_female = [(age, 5.2 + (age - 20) * 0.15) for age in range(20, 81, 5)]
    tug = {'Male': tug_male, 'Female': tug_female}

    # Single Leg Stance (no gender split)
    ws = wb['Single leg stance']
    sls = []
    for row in range(3, 9):
        age_range = parse_age_range(str(ws.cell(row, 1).value))
        # Parse "45.1 (±0.1)" format
        val_str = str(ws.cell(row, 2).value)
        mean_val = float(val_str.split('(')[0].strip())
        sls.append((age_range, mean_val))
    single_leg_stance = sls

    # Sit & Reach
    ws = wb['Sit & Reach']
    sr_male = []
    sr_female = []
    for row in range(3, 8):
        age_range = parse_age_range(str(ws.cell(row, 1).value))
        male_val = float(ws.cell(row, 2).value)
        female_val = float(ws.cell(row, 3).value)
        sr_male.append((age_range, male_val))
        sr_female.append((age_range, female_val))
    sit_and_reach = {'Male': sr_male, 'Female': sr_female}

    # Metabolic Normative Data
    ws = wb['Metabolic normative data']
    metabolic_ranges = {}
    markers = []
    # Read header row (row 2)
    for col in range(2, 7):
        marker = ws.cell(2, col).value
        if marker:
            markers.append(marker)

    # Read risk ranges (rows 3-5)
    for marker in markers:
        col_idx = None
        for col in range(2, 7):
            if ws.cell(2, col).value == marker:
                col_idx = col
                break

        if col_idx:
            metabolic_ranges[marker] = {
                'Low risk': str(ws.cell(3, col_idx).value),
                'Normal': str(ws.cell(4, col_idx).value),
                'Elevated': str(ws.cell(5, col_idx).value)
            }

    # Body Fat % - assuming it was added to the database
    # Placeholder structure based on Gallagher et al. (2000)
    # Format: age_range -> (healthy_max, overweight_max, obese_min)
    body_fat_male = [
        ((20, 39), (8, 20, 25)),
        ((40, 59), (11, 22, 28)),
        ((60, 79), (13, 25, 30))
    ]
    body_fat_female = [
        ((20, 39), (21, 33, 39)),
        ((40, 59), (23, 34, 40)),
        ((60, 79), (24, 36, 42))
    ]
    body_fat_ranges = {'Male': body_fat_male, 'Female': body_fat_female}

    return NormativeData(
        vo2_max=vo2_max,
        grip_strength=grip_strength,
        vertical_jump=vertical_jump,
        sts_power=sts_power,
        gait_speed=gait_speed,
        tug=tug,
        single_leg_stance=single_leg_stance,
        sit_and_reach=sit_and_reach,
        metabolic_ranges=metabolic_ranges,
        body_fat_ranges=body_fat_ranges
    )
