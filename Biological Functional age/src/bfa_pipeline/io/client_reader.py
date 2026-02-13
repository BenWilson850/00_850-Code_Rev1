"""Reader for client test data from Excel workbooks."""

from __future__ import annotations
from pathlib import Path
from typing import List, Optional
import openpyxl

from ..core.test_scoring import TestResults


# Fixed row positions for template
ROW_GENDER = 1
ROW_NAME = 2
ROW_AGE = 4
ROW_VO2_MAX = 5
ROW_FEV1 = 6
ROW_GRIP = 7
ROW_STS = 8
ROW_VJ = 9
ROW_BODY_FAT = 10
ROW_WHTR = 11
ROW_GLUCOSE = 12
ROW_HBA1C = 13
ROW_HOMA = 14
ROW_APOB = 15
ROW_HSCRP = 16
ROW_GAIT = 17
ROW_TUG = 18
ROW_SLS = 19
ROW_SIT_REACH = 20
ROW_PROCESSING = 21
ROW_MEMORY = 22

# Column for current values
COL_VALUE = 2  # Column B


def safe_float(value) -> Optional[float]:
    """
    Safely convert a value to float, returning None if not possible.

    Args:
        value: Value to convert

    Returns:
        Float value or None
    """
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_gender(gender_str) -> str:
    """
    Parse gender string to 'Male' or 'Female'.

    Args:
        gender_str: Gender string from Excel

    Returns:
        'Male' or 'Female'
    """
    if gender_str is None:
        return 'Male'  # Default

    gender_lower = str(gender_str).strip().lower()

    if 'male' in gender_lower and 'female' not in gender_lower:
        return 'Male'
    elif 'female' in gender_lower:
        return 'Female'
    else:
        return 'Male'  # Default


def read_client_sheet(ws) -> TestResults:
    """
    Read a single client sheet from the workbook.

    Args:
        ws: Worksheet object

    Returns:
        TestResults for this client
    """
    # Read metadata
    gender_raw = ws.cell(ROW_GENDER, COL_VALUE).value
    gender = parse_gender(gender_raw)

    name = ws.cell(ROW_NAME, COL_VALUE).value
    if name is None:
        name = ws.title  # Use sheet name if no name provided

    age_raw = ws.cell(ROW_AGE, COL_VALUE).value
    age = safe_float(age_raw)
    if age is None:
        raise ValueError(f"Age missing or invalid for client: {name}")

    # Read test results
    return TestResults(
        name=str(name).strip(),
        age=age,
        gender=gender,
        # Vitality
        vo2_max=safe_float(ws.cell(ROW_VO2_MAX, COL_VALUE).value),
        fev1=safe_float(ws.cell(ROW_FEV1, COL_VALUE).value),
        # Strength
        grip_strength=safe_float(ws.cell(ROW_GRIP, COL_VALUE).value),
        sts_power=safe_float(ws.cell(ROW_STS, COL_VALUE).value),
        vertical_jump=safe_float(ws.cell(ROW_VJ, COL_VALUE).value),
        # Metabolic
        body_fat_pct=safe_float(ws.cell(ROW_BODY_FAT, COL_VALUE).value),
        whtr=safe_float(ws.cell(ROW_WHTR, COL_VALUE).value),
        fasting_glucose=safe_float(ws.cell(ROW_GLUCOSE, COL_VALUE).value),
        hba1c=safe_float(ws.cell(ROW_HBA1C, COL_VALUE).value),
        homa_ir=safe_float(ws.cell(ROW_HOMA, COL_VALUE).value),
        apob=safe_float(ws.cell(ROW_APOB, COL_VALUE).value),
        hscrp=safe_float(ws.cell(ROW_HSCRP, COL_VALUE).value),
        # Mobility
        gait_speed=safe_float(ws.cell(ROW_GAIT, COL_VALUE).value),
        tug=safe_float(ws.cell(ROW_TUG, COL_VALUE).value),
        single_leg_stance=safe_float(ws.cell(ROW_SLS, COL_VALUE).value),
        sit_and_reach=safe_float(ws.cell(ROW_SIT_REACH, COL_VALUE).value),
        # Cognitive
        processing_speed=safe_float(ws.cell(ROW_PROCESSING, COL_VALUE).value),
        working_memory=safe_float(ws.cell(ROW_MEMORY, COL_VALUE).value),
    )


def read_client_workbook(workbook_path: str | Path) -> List[TestResults]:
    """
    Read all client sheets from a workbook.

    Args:
        workbook_path: Path to Excel workbook

    Returns:
        List of TestResults, one per sheet
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Client workbook not found: {workbook_path}")

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    clients = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            client = read_client_sheet(ws)
            clients.append(client)
        except Exception as e:
            print(f"Warning: Failed to read sheet '{sheet_name}': {e}")
            continue

    if not clients:
        raise ValueError(f"No valid client data found in {workbook_path}")

    return clients
